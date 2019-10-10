import bs4 as bs
import urllib.request
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import os.path
from datetime import date

flightData = []



def scrapeForFlights():

    #Flight from ROA to wester US < $450
    urlWestUSSpringBreak='https://www.google.com/flights?lite=0#flt=ROA..2020-03-07*.ROA.2020-03-13;c:USD;e:1;p:45000.2.USD;sd:1;er:185569991.1715776367.655103616.-913911133;t:e'

    # user-agent is necessary to get past amazon's non-browser traffic block.
    user_agent = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.1916.47 Safari/537.36'
    request = urllib.request.Request(urlWestUSSpringBreak,headers={'User-Agent': user_agent})

    try:
        source = urllib.request.urlopen(request).read()
    except :
        print("Bad URL:" + urlWestUSSpringBreak)
        return
    soup = bs.BeautifulSoup(source,'lxml')

    print(soup.find('p'))

    flights = soup.select('.uKOpFp4SF2X__info-container')
    print(flights)
    if (len(flights) > 0) :
        for flight in flights:
            flightName = flight.select('h3[class="flt-subhead1"]')[0].text
            # Google hides the href de-encrypts href in js, need to search or find out how to interpret the encyption
            # flightUrl = flight.get('href')
            # flightId = flightUrl.split('/',10)[4]

            # try:
            flightPrice = float(flight.select('.uKOpFp4SF2X__price flt-subhead2')[0].text.split('$',2)[1].strip())
            # except:
                # continue
            flightData.append([flightName,flightPrice])

    else :
        print("Scrape complete.")
        updateSpreadsheet()
        return;

def updateSpreadsheet():

    ##Sub functions
    def findFirstEmptyRow():
        for cell in sheet["A"]:
            if cell.value is None :
                return cell.row
        return len(sheet["A"])+1


    filename = "flight-data.xlsx"

    #initalize spreadsheet
    if os.path.exists(filename) :
        workbook = load_workbook(filename=filename)
        sheet = workbook.active
    else :
        workbook = Workbook()
        sheet = workbook.active

        # Setting the headers of the data sheet
        sheet["A1"] = "Destinantion"
        sheet["B1"] = "Price"
        sheet["C1"] = "Departure Date"
        sheet["D1"] = "Return Date"
        sheet["D1"] = "Source URL"
        # sheet.column_dimensions["D"].number_format = '$0.00'

    columnToday=len(sheet[1])+1
    # sheet.column_dimensions[get_column_letter(columnToday)].number_format = '$0.00'
    # sheet.cell(row = 1, column= columnToday).number_format = '%Y-%m-%d'
    sheet.cell(row = 1, column= columnToday).value = date.today()

    for flight in flightData:
        flightRow = findFirstEmptyRow()
        sheet.cell(row=flightRow, column=1).value = flight[0]
        sheet.cell(row=flightRow, column=2).value = flight[1]


    # sheet["D"].number_format = '$0.00''

    workbook.save(filename=filename)

scrapeForFlights()
